# coding=utf-8
import openpyxl

file_name = r"./工伤赔偿计算公式.xlsx"
olw = openpyxl.load_workbook(file_name)
sheet_names = olw.get_sheet_names()
all = []
for name in sheet_names:
    ws = olw.get_sheet_by_name(name)
    container = []
    for index in range(4, 14):
        item = {
            ws.cell(index, 1).value: {"一次性医疗补助金": ws.cell(index, 7).value, "一次性就业补助金": ws.cell(index, 8).value}
        }
        container.append(item)
    is_PW = False
    is_SW = False
    for item in container:
        (key, value), = item.items()
        一次性医疗补助金 = value["一次性医疗补助金"];
        一次性就业补助金 = value["一次性就业补助金"];
        if not is_SW:
            if 一次性医疗补助金 and isinstance(一次性医疗补助金, str) and "PW" in 一次性医疗补助金:
                is_PW = True
            elif 一次性就业补助金 and isinstance(一次性就业补助金, str) and "PW" in 一次性就业补助金:
                is_PW = True
        if not is_SW:
            if isinstance(一次性医疗补助金, str) and ((一次性医疗补助金 and "SW" in 一次性医疗补助金) or (一次性就业补助金 and "SW" in 一次性就业补助金)):
                is_SW = True
    if is_PW and is_SW:
        raise RuntimeError("错误了.")
    # if not is_PW and not is_SW:
    #     raise RuntimeError("都没有匹配错误了.")
    ret = {"name": name, "is_SW": is_SW, "is_PW": is_PW, "container": container, }
    all.append(ret)

DEAL_DATA = ['黑龙江', '吉林', '辽宁', '内蒙古', '宁夏', '青海', '山东', '山西', '陕西', '四川', '西藏', '新疆', '云南', '浙江']

for data in all:
    name = data["name"]
    if name not in DEAL_DATA:
        continue
    is_SW = data["is_SW"]
    is_PW = data["is_PW"]
    grantBaseType = "不知道"
    if is_SW:
        grantBaseType = 1
    if is_PW:
        grantBaseType = 2
    for _it in data["container"]:
        for (key, value) in _it.items():
            grantMedical = 0
            grantWork = 0
            一次性医疗补助金 = value["一次性医疗补助金"]
            一次性就业补助金 = value["一次性就业补助金"]
            if is_SW:
                if 一次性医疗补助金 and isinstance(一次性医疗补助金, str):
                    grantMedical = 一次性医疗补助金.replace("SW/12*", "").strip()
                if 一次性就业补助金 and isinstance(一次性就业补助金, str):
                    grantWork = 一次性就业补助金.replace("SW/12*", "").strip()
            if is_PW:
                if 一次性医疗补助金 and isinstance(一次性医疗补助金, str):
                    grantMedical = 一次性医疗补助金.replace("PW*", "").strip()
                if 一次性就业补助金 and isinstance(一次性就业补助金, str):
                    grantWork = 一次性就业补助金.replace("PW*", "").strip()
            sql = "UPDATE tool_injuryitem SET grantBaseType={},grantMedical={},grantWork={} WHERE areaName='{}' and rangeName='{}';".format(
                grantBaseType,
                grantMedical,
                grantWork,
                name,
                key, )
            print(sql)
